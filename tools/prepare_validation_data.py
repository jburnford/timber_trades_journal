#!/usr/bin/env python3
"""
Prepare validation datasets for comparing human vs automated transcription.

This script:
1. Exports human ground truth from Excel (1883 and 1889 London imports)
2. Filters automated data to matching scope (year + London destination)
3. Adds hybrid_arrival_date field to automated data
4. Outputs aligned CSVs ready for comparison
"""

import csv
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple


def export_human_ground_truth_1883(excel_path: Path, output_path: Path) -> int:
    """
    Export 1883 London imports from Excel to CSV.

    Returns:
        Number of rows exported
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = wb['London imports 1883']

    rows_exported = 0
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header
        writer.writerow(['date', 'origin_port', 'quantity', 'unit', 'product'])

        # Process data rows (skip header row)
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 1:  # Skip header
                continue

            date, origin_port, quantity, unit, product = row[0], row[1], row[2], row[3], row[4]

            # Skip empty rows
            if not date and not origin_port:
                continue

            # Normalize date to YYYY-MM-DD format
            if isinstance(date, datetime):
                date_str = date.strftime('%Y-%m-%d')
            elif isinstance(date, str):
                date_str = date  # Assume already formatted
            else:
                date_str = str(date) if date else ''

            writer.writerow([date_str, origin_port, quantity, unit, product])
            rows_exported += 1

    wb.close()
    print(f"✓ Exported {rows_exported:,} rows from 1883 sheet")
    return rows_exported


def export_human_ground_truth_1889(excel_path: Path, output_path: Path) -> int:
    """
    Export 1889 London imports from Excel to CSV.

    Returns:
        Number of rows exported
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet = wb['Lon imp. 1889']

    rows_exported = 0
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header - note additional fields
        writer.writerow(['date', 'dock', 'origin_port', 'quantity', 'unit', 'products', 'merchants'])

        # Process data rows (skip header row)
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 1:  # Skip header
                continue

            date, dock, origin_port, quantity, unit, products, merchants = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            )

            # Skip empty rows
            if not date and not origin_port:
                continue

            # Normalize date to YYYY-MM-DD format
            if isinstance(date, datetime):
                date_str = date.strftime('%Y-%m-%d')
            elif isinstance(date, str):
                date_str = date  # Assume already formatted
            else:
                date_str = str(date) if date else ''

            writer.writerow([date_str, dock, origin_port, quantity, unit, products, merchants])
            rows_exported += 1

    wb.close()
    print(f"✓ Exported {rows_exported:,} rows from 1889 sheet")
    return rows_exported


def create_hybrid_date(arrival_day: str, arrival_month: str, arrival_year: str,
                       pub_day: str, pub_month: str, pub_year: str) -> str:
    """
    Create hybrid arrival date: use arrival date if available, else publication date.

    Returns:
        Date in YYYY-MM-DD format
    """
    # Try to use arrival date
    if arrival_day and arrival_month and arrival_year:
        try:
            # Convert month name to number
            month_map = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12,
                'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'Jun': 6,
                'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
            }
            month_num = month_map.get(arrival_month, None)
            if month_num:
                return f"{arrival_year}-{month_num:02d}-{int(arrival_day):02d}"
        except (ValueError, TypeError):
            pass

    # Fall back to publication date
    if pub_day and pub_month and pub_year:
        try:
            month_map = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12,
                'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'Jun': 6,
                'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
            }
            month_num = month_map.get(pub_month, None)
            if month_num:
                return f"{pub_year}-{month_num:02d}-{int(pub_day):02d}"
        except (ValueError, TypeError):
            pass

    return ''


def filter_automated_cargo_data(input_csv: Path, output_csv: Path, year: int,
                                 destination_filter: str = 'London') -> Tuple[int, int, int]:
    """
    Filter automated cargo data to specific year and London destinations.
    Add hybrid_arrival_date field.

    Returns:
        (total_rows, filtered_rows, rows_with_arrival_date)
    """
    csv.field_size_limit(1000000)  # Handle large cargo fields

    total_rows = 0
    filtered_rows = 0
    rows_with_arrival_date = 0

    with open(input_csv, 'r', encoding='utf-8') as f_in:
        reader = csv.DictReader(f_in)

        # Get field names and add hybrid_date
        fieldnames = reader.fieldnames + ['hybrid_arrival_date', 'date_source']

        with open(output_csv, 'w', newline='', encoding='utf-8') as f_out:
            writer = csv.DictWriter(f_out, fieldnames=fieldnames)
            writer.writeheader()

            for row in reader:
                total_rows += 1

                # Check year (from record_id lookup or direct year field)
                # We'll need to join with shipments to get year - for now use record_id
                # This is a simplification - may need refinement

                # For now, we'll process all rows and add hybrid dates
                # User will need to filter by year after joining with shipments table

                # Create hybrid date
                hybrid_date = create_hybrid_date(
                    row.get('arrival_day', ''),
                    row.get('arrival_month', ''),
                    row.get('arrival_year', ''),
                    row.get('publication_day', ''),
                    row.get('publication_month', ''),
                    row.get('publication_year', '')
                )

                # Determine date source
                if row.get('arrival_day') and row.get('arrival_month') and row.get('arrival_year'):
                    date_source = 'arrival'
                    rows_with_arrival_date += 1
                else:
                    date_source = 'publication'

                row['hybrid_arrival_date'] = hybrid_date
                row['date_source'] = date_source

                writer.writerow(row)
                filtered_rows += 1

    return total_rows, filtered_rows, rows_with_arrival_date


def filter_automated_shipments_for_year(shipments_csv: Path, cargo_csv: Path,
                                         output_csv: Path, year: int) -> int:
    """
    Join cargo details with shipments, filter by year and London destination.
    Creates complete cargo records with all shipment fields included.

    Returns:
        Number of cargo rows for specified year + London
    """
    csv.field_size_limit(1000000)

    # First, load ALL shipments into memory indexed by source_file + line_number
    # This is the proper join key between cargo and shipments
    shipment_lookup = {}
    with open(shipments_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Create composite key: source_file + line_number
            key = (row['source_file'], row['line_number'])
            shipment_lookup[key] = row

    print(f"  Loaded {len(shipment_lookup):,} shipment records for lookup")

    # Now join cargo with shipments and filter
    filtered_rows = 0
    with open(cargo_csv, 'r', encoding='utf-8') as f_in:
        reader = csv.DictReader(f_in)

        # Create combined fieldnames: cargo fields + shipment fields
        cargo_fields = reader.fieldnames
        sample_shipment = next(iter(shipment_lookup.values())) if shipment_lookup else {}
        shipment_fields = list(sample_shipment.keys())

        # Combined fields (avoid duplicates)
        all_fields = list(cargo_fields)
        for field in shipment_fields:
            if field not in all_fields and field not in ['source_file', 'line_number']:
                all_fields.append(field)

        with open(output_csv, 'w', newline='', encoding='utf-8') as f_out:
            writer = csv.DictWriter(f_out, fieldnames=all_fields)
            writer.writeheader()

            for cargo_row in reader:
                # Join with shipments using source_file + line_number
                key = (cargo_row['source_file'], cargo_row['line_number'])

                if key in shipment_lookup:
                    shipment_row = shipment_lookup[key]

                    # Check year and London destination
                    pub_year = shipment_row.get('publication_year', '')
                    destination = shipment_row.get('destination_port', '').upper()

                    # Debug: show first few checks
                    if filtered_rows < 5:
                        print(f"    Checking: year={pub_year} (target={year}), dest={destination}")

                    if pub_year == str(year) and 'LONDON' in destination:
                        # Merge rows
                        combined_row = {**cargo_row, **shipment_row}
                        writer.writerow(combined_row)
                        filtered_rows += 1
                        if filtered_rows == 1:
                            print(f"    First match found!")

    return filtered_rows


def main():
    print("=" * 80)
    print("PREPARING VALIDATION DATASETS")
    print("=" * 80)

    # Paths
    base_dir = Path("/home/jic823/TTJ Forest of Numbers")
    excel_file = base_dir / "London Timber imports data ttj.xlsx"
    output_dir = base_dir / "validation"
    output_dir.mkdir(exist_ok=True)

    # Automated data paths (use original, not normalized, to match source files)
    cargo_details = base_dir / "final_output/ttj_cargo_details.csv"
    shipments = base_dir / "parsed_output/ttj_shipments_multipage.csv"

    print(f"\nInput files:")
    print(f"  Human data: {excel_file}")
    print(f"  Automated cargo: {cargo_details}")
    print(f"  Automated shipments: {shipments}")
    print(f"\nOutput directory: {output_dir}")
    print()

    # Export human ground truth
    print("Step 1: Exporting human ground truth data...")
    print("-" * 80)

    gt_1883 = output_dir / "ground_truth_1883_london.csv"
    rows_1883 = export_human_ground_truth_1883(excel_file, gt_1883)

    gt_1889 = output_dir / "ground_truth_1889_london.csv"
    rows_1889 = export_human_ground_truth_1889(excel_file, gt_1889)

    # Filter for 1883 + London (joins cargo with shipments)
    print("\nStep 2: Filtering and joining automated data for 1883 London imports...")
    print("-" * 80)

    auto_1883 = output_dir / "automated_1883_london.csv"
    filtered_1883 = filter_automated_shipments_for_year(
        shipments, cargo_details, auto_1883, 1883
    )
    print(f"✓ Extracted {filtered_1883:,} cargo records for 1883 London")

    # Filter for 1889 + London
    print("\nStep 3: Filtering and joining automated data for 1889 London imports...")
    print("-" * 80)

    auto_1889 = output_dir / "automated_1889_london.csv"
    filtered_1889 = filter_automated_shipments_for_year(
        shipments, cargo_details, auto_1889, 1889
    )
    print(f"✓ Extracted {filtered_1889:,} cargo records for 1889 London")

    # Now add hybrid dates to both
    print("\nStep 4: Adding hybrid arrival dates...")
    print("-" * 80)

    # Count date sources for both years
    for year_file, year_name in [(auto_1883, "1883"), (auto_1889, "1889")]:
        with_arrival = 0
        total = 0
        with open(year_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                total += 1
                if row.get('arrival_day') and row.get('arrival_month') and row.get('arrival_year'):
                    with_arrival += 1

        if total > 0:
            print(f"  {year_name}: {with_arrival}/{total} with arrival dates ({100*with_arrival/total:.1f}%)")
        else:
            print(f"  {year_name}: No records found")

    # Summary
    print("\n" + "=" * 80)
    print("DATA PREPARATION COMPLETE")
    print("=" * 80)
    print("\nGround Truth Data:")
    print(f"  1883: {rows_1883:,} cargo records → {gt_1883}")
    print(f"  1889: {rows_1889:,} cargo records → {gt_1889}")
    print("\nAutomated Data:")
    print(f"  1883: {filtered_1883:,} cargo records → {auto_1883}")
    print(f"  1889: {filtered_1889:,} cargo records → {auto_1889}")
    print("\nNext step: Run match_cargo_records.py to create matched pairs")
    print("=" * 80)


if __name__ == '__main__':
    main()
